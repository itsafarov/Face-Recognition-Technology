"""
Генератор профессиональных отчетов
"""

import os
import json
import datetime
import html
import time
from typing import List, Dict, Any, Tuple, Optional
from collections import defaultdict

from core.config import Config

# Удален импорт FaceRecord и ProcessingMetrics - используем отложенные импорты

class ReportGenerator:
    """Генератор профессиональных отчетов"""
    
    def __init__(self, output_dir: str):
        self.output_dir = output_dir
        self.reports_dir = os.path.join(output_dir, Config.REPORTS_FOLDER)
        self.images_dir = os.path.join(output_dir, Config.IMAGE_FOLDER)
        
        # Создаем папку для отчетов если не существует
        os.makedirs(self.reports_dir, exist_ok=True)
    
    def generate_html_report(self, records: List, metrics) -> str:
        """Генерация HTML отчета с встроенными фото"""
        print(f"📊 Анализ статистики для HTML отчета...")
        
        # Используем отложенный импорт для избежания циклической зависимости
        from core.statistics import StatisticsAnalyzer
        from core.models import FaceRecord, ProcessingMetrics
        
        # Проверяем типы
        if records and isinstance(records[0], dict):
            # Если переданы словари, конвертируем в FaceRecord
            face_records = [FaceRecord(**record) for record in records]
        else:
            # Уже FaceRecord объекты
            face_records = records
        
        stats = StatisticsAnalyzer.analyze(face_records)
        print(f"📊 Создание HTML отчета из {len(face_records)} записей...")
        
        try:
            html_content = self._build_html_template(face_records, metrics, stats)
            report_path = os.path.join(self.reports_dir, Config.HTML_REPORT)
            
            with open(report_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            print(f"✅ HTML отчет создан: {report_path}")
            return report_path
            
        except Exception as e:
            print(f"❌ Ошибка создания HTML отчета: {e}")
            import traceback
            traceback.print_exc()
            # Создаем простой HTML отчет в случае ошибки
            return self._create_fallback_html_report(face_records, metrics)
    
    def _build_html_template(self, records: List, metrics, stats: Dict) -> str:
        """Шаблон HTML отчета с улучшенной статистикой"""
        # Генерация строк таблицы
        print("📋 Генерация строк таблицы...")
        rows_html = ""
        for i, record in enumerate(records):
            rows_html += record.to_html_row(i)
        
        print("📊 Подготовка статистики...")
        # Время и метрики
        elapsed = metrics.elapsed_time
        speed = metrics.total_records / elapsed if elapsed > 0 else 0
        
        # Подготовка данных для JavaScript
        # Преобразуем ключи и значения в строки для JSON
        company_labels = []
        company_data = []
        for k, v in stats['by_company'].items():
            company_labels.append(str(k))
            company_data.append(v)
        
        gender_labels = []
        gender_data = []
        for k, v in stats['by_gender'].items():
            gender_labels.append(str(k))
            gender_data.append(v)
        
        # Ограничиваем количество элементов для графиков
        max_chart_items = 15
        if len(company_labels) > max_chart_items:
            # Берем топ N компаний
            combined = list(zip(company_labels, company_data))
            combined.sort(key=lambda x: x[1], reverse=True)
            company_labels = [str(x[0]) for x in combined[:max_chart_items]]
            company_data = [x[1] for x in combined[:max_chart_items]]
        
        # Преобразуем в JSON строки с экранированием
        company_labels_json = json.dumps(company_labels, ensure_ascii=False)
        company_data_json = json.dumps(company_data, ensure_ascii=False)
        gender_labels_json = json.dumps(gender_labels, ensure_ascii=False)
        gender_data_json = json.dumps(gender_data, ensure_ascii=False)
        
        # Получаем HTML статистики
        stats_html = self._generate_stats_html(stats, metrics)
        
        # Получаем опции фильтра
        filter_options = self._generate_filter_options('company', records)
        
        # Генерация даты и времени
        current_datetime = datetime.datetime.now()
        
        print("🎨 Формирование HTML отчета...")
        # Формируем полный HTML
        return self._get_full_html_template(
            current_datetime=current_datetime,
            metrics=metrics,
            speed=speed,
            company_labels_json=company_labels_json,
            company_data_json=company_data_json,
            gender_labels_json=gender_labels_json,
            gender_data_json=gender_data_json,
            rows_html=rows_html,
            elapsed=elapsed,
            stats_html=stats_html,
            filter_options=filter_options,
            total_records=len(records)
        )
    
    def _get_full_html_template(self, **kwargs) -> str:
        """Возвращает полный HTML шаблон"""
        # Основной стиль CSS
        style_content = self._get_css_styles()
        
        # JavaScript код с исправленным парсингом JSON
        javascript_content = self._get_javascript_code()
        
        # Форматируем дату
        date_str = kwargs['current_datetime'].strftime("%d.%m.%Y %H:%M")
        
        return f'''<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Отчет по распознаванию лиц</title>
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <style>
        {style_content}
    </style>
</head>
<body>
    <div class="container">
        <header class="header">
            <h1><i class="fas fa-id-card"></i> Отчет по распознаванию лиц</h1>
            <div style="display: flex; justify-content: center; gap: 30px; flex-wrap: wrap;">
                <div style="background: rgba(255,255,255,0.1); padding: 10px 20px; border-radius: 50px;">
                    <i class="fas fa-calendar"></i> {date_str}
                </div>
                <div style="background: rgba(255,255,255,0.1); padding: 10px 20px; border-radius: 50px;">
                    <i class="fas fa-database"></i> {kwargs['metrics'].total_records:,} записей
                </div>
            </div>
        </header>
        
        <div class="summary-cards">
            <div class="card card-success">
                <i class="fas fa-check-circle"></i>
                <div>
                    <h3 style="font-size: 1.8rem; margin: 0;">{kwargs['metrics'].valid_images:,}</h3>
                    <p style="margin: 0; font-size: 14px;">Успешных фото</p>
                </div>
            </div>
            <div class="card card-warning">
                <i class="fas fa-times-circle"></i>
                <div>
                    <h3 style="font-size: 1.8rem; margin: 0;">{kwargs['metrics'].failed_images:,}</h3>
                    <p style="margin: 0; font-size: 14px;">Ошибок загрузки</p>
                </div>
            </div>
            <div class="card card-info">
                <i class="fas fa-bolt"></i>
                <div>
                    <h3 style="font-size: 1.8rem; margin: 0;">{kwargs['speed']:.0f}/сек</h3>
                    <p style="margin: 0; font-size: 14px;">Скорость</p>
                </div>
            </div>
            <div class="card card-primary">
                <i class="fas fa-chart-line"></i>
                <div>
                    <h3 style="font-size: 1.8rem; margin: 0;">{kwargs['metrics'].success_rate:.1f}%</h3>
                    <p style="margin: 0; font-size: 14px;">Успешность</p>
                </div>
            </div>
        </div>
        
        <div class="statistics-section">
            <h2 style="color: #2c3e50; margin-bottom: 20px; display: flex; align-items: center; gap: 10px;">
                <i class="fas fa-chart-bar"></i> Статистика анализа
            </h2>
            
            {kwargs['stats_html']}
        </div>
        
        <div class="table-container">
            <h2 style="color: #2c3e50; margin-bottom: 20px; display: flex; align-items: center; gap: 10px;">
                <i class="fas fa-table"></i> Детальные данные ({kwargs['total_records']:,} записей)
            </h2>
            
            <div class="filters">
                <div class="filter-group">
                    <label><i class="fas fa-filter"></i> Компания:</label>
                    <select id="companyFilter" class="filter-select" onchange="filterTable()">
                        <option value="">Все компании</option>
                        {kwargs['filter_options']}
                    </select>
                </div>
                <div class="filter-group">
                    <label>Тип события:</label>
                    <select id="eventFilter" class="filter-select" onchange="filterTable()">
                        <option value="">Все события</option>
                        <option value="1">Распознавание</option>
                        <option value="2">Событие</option>
                    </select>
                </div>
                <div class="filter-group">
                    <label>Статус списка:</label>
                    <select id="listFilter" class="filter-select" onchange="filterTable()">
                        <option value="">Все</option>
                        <option value="1">В списке</option>
                        <option value="0">Не в списке</option>
                    </select>
                </div>
                <button class="filter-btn" onclick="resetFilters()">
                    <i class="fas fa-redo"></i> Сбросить
                </button>
            </div>
            
            <div class="table-wrapper">
                <table id="dataTable">
                    <thead>
                        <tr>
                            <th><i class="fas fa-hashtag"></i> №</th>
                            <th><i class="fas fa-clock"></i> Время</th>
                            <th><i class="fas fa-microchip"></i> Устройство</th>
                            <th><i class="fas fa-user"></i> Пользователь</th>
                            <th><i class="fas fa-venus-mars"></i> Пол</th>
                            <th><i class="fas fa-birthday-cake"></i> Возраст</th>
                            <th><i class="fas fa-percentage"></i> Совпадение</th>
                            <th><i class="fas fa-id-badge"></i> ID Лица</th>
                            <th><i class="fas fa-building"></i> Компания</th>
                            <th><i class="fas fa-calendar-check"></i> Тип</th>
                            <th><i class="fas fa-list-check"></i> Список</th>
                            <th><i class="fas fa-image"></i> Фото</th>
                        </tr>
                    </thead>
                    <tbody>
                        {kwargs['rows_html']}
                    </tbody>
                </table>
            </div>
            
            <div class="actions">
                <button onclick="window.print()" class="btn btn-print">
                    <i class="fas fa-print"></i> Печать
                </button>
                <a href="#" onclick="saveAsPDF()" class="btn btn-pdf">
                    <i class="fas fa-file-pdf"></i> Сохранить как PDF
                </a>
                <a href="#" onclick="exportToExcel()" class="btn btn-excel">
                    <i class="fas fa-file-excel"></i> Экспорт в Excel
                </a>
            </div>
        </div>
        
        <footer class="footer">
            <p><i class="fas fa-info-circle"></i> Отчет сгенерирован автоматически</p>
            <p style="margin-top: 10px; font-size: 14px;">
                Время обработки: {kwargs['elapsed']:.1f} сек. | Кэшировано фото: {kwargs['metrics'].cached_images:,} |
                Уникальных пользователей: {len(kwargs['metrics'].unique_users):,} |
                Устройств: {len(kwargs['metrics'].unique_devices):,} |
                Компаний: {len(kwargs['metrics'].unique_companies):,}
            </p>
        </footer>
    </div>
    
    <!-- Модальное окно для предпросмотра фото -->
    <div id="imageModal" class="modal" onclick="closeModal()">
        <span class="close">&times;</span>
        <img class="modal-content" id="modalImage">
        <div id="modalInfo" class="modal-info"></div>
    </div>
    
    <script>
        // Данные для графиков (исправлено - не парсим строку, а используем переменные)
        const companyLabels = {kwargs['company_labels_json']};
        const companyData = {kwargs['company_data_json']};
        const genderLabels = {kwargs['gender_labels_json']};
        const genderData = {kwargs['gender_data_json']};
        
        {javascript_content}
    </script>
    
    <!-- Подключаем библиотеку для создания PDF -->
    <script src="https://cdnjs.cloudflare.com/ajax/libs/html2pdf.js/0.10.1/html2pdf.bundle.min.js"></script>
</body>
</html>'''
    
    def _get_css_styles(self) -> str:
        """Возвращает CSS стили"""
        return '''* {
    margin: 0;
    padding: 0;
    box-sizing: border-box;
}

body {
    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    min-height: 100vh;
    padding: 20px;
}

.container {
    max-width: 95%;
    margin: 0 auto;
    background: white;
    border-radius: 20px;
    box-shadow: 0 20px 60px rgba(0,0,0,0.3);
    overflow: hidden;
}

.header {
    background: linear-gradient(135deg, #2c3e50 0%, #4a6491 100%);
    color: white;
    padding: 30px;
    text-align: center;
}

.header h1 {
    font-size: 2.5rem;
    margin-bottom: 20px;
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 15px;
}

.summary-cards {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
    gap: 20px;
    padding: 25px;
    background: #f8f9fa;
}

.card {
    display: flex;
    align-items: center;
    gap: 15px;
    padding: 20px;
    border-radius: 12px;
    color: white;
    transition: transform 0.3s ease;
}

.card:hover {
    transform: translateY(-5px);
}

.card i {
    font-size: 2rem;
}

.card-success { background: linear-gradient(135deg, #00b09b 0%, #96c93d 100%); }
.card-warning { background: linear-gradient(135deg, #f46b45 0%, #eea849 100%); }
.card-info { background: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%); }
.card-primary { background: linear-gradient(135deg, #a18cd1 0%, #fbc2eb 100%); }

.statistics-section {
    padding: 25px;
    background: white;
}

.stats-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
    gap: 20px;
    margin-top: 20px;
}

.stat-box {
    background: white;
    border-radius: 12px;
    padding: 20px;
    box-shadow: 0 5px 15px rgba(0,0,0,0.1);
    border: 1px solid #e0e0e0;
}

.stat-box h3 {
    color: #2c3e50;
    margin-bottom: 15px;
    padding-bottom: 10px;
    border-bottom: 2px solid #4a6fa5;
    display: flex;
    align-items: center;
    gap: 10px;
}

.stat-item {
    display: flex;
    justify-content: space-between;
    padding: 8px 0;
    border-bottom: 1px dashed #eee;
}

.stat-item:last-child {
    border-bottom: none;
}

.stat-label {
    color: #666;
}

.stat-value {
    font-weight: bold;
    color: #2c3e50;
}

.chart-container {
    height: 300px;
    margin-top: 15px;
}

.table-container {
    padding: 25px;
}

.table-wrapper {
    overflow-x: auto;
    border-radius: 12px;
    border: 1px solid #e0e0e0;
    margin-top: 20px;
    max-height: 70vh;
    overflow-y: auto;
}

table {
    width: 100%;
    border-collapse: collapse;
    min-width: 1400px;
}

thead {
    background: linear-gradient(135deg, #4a6fa5 0%, #2c3e50 100%);
    color: white;
    position: sticky;
    top: 0;
    z-index: 10;
}

th {
    padding: 15px;
    text-align: left;
    font-weight: 500;
    font-size: 14px;
    position: sticky;
    top: 0;
    background: inherit;
}

th i {
    margin-right: 8px;
}

tbody tr:hover {
    background-color: #f5f9ff !important;
}

.footer {
    background: #f8f9fa;
    padding: 20px 30px;
    border-top: 1px solid #e0e0e0;
    text-align: center;
    color: #666;
}

.actions {
    display: flex;
    gap: 15px;
    justify-content: center;
    margin-top: 20px;
    flex-wrap: wrap;
}

.btn {
    padding: 12px 24px;
    border: none;
    border-radius: 8px;
    font-weight: 500;
    cursor: pointer;
    display: inline-flex;
    align-items: center;
    gap: 8px;
    transition: all 0.3s ease;
    text-decoration: none;
    color: white;
    font-size: 14px;
}

.btn-pdf {
    background: linear-gradient(135deg, #ff6b6b 0%, #ee5a52 100%);
}

.btn-excel {
    background: linear-gradient(135deg, #21d160 0%, #2bb673 100%);
}

.btn-print {
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
}

.btn:hover {
    opacity: 0.9;
    transform: translateY(-2px);
    box-shadow: 0 5px 15px rgba(0,0,0,0.2);
}

/* Модальное окно для предпросмотра фото */
.modal {
    display: none;
    position: fixed;
    z-index: 1000;
    left: 0;
    top: 0;
    width: 100%;
    height: 100%;
    background-color: rgba(0,0,0,0.9);
    justify-content: center;
    align-items: center;
}

.modal-content {
    max-width: 90%;
    max-height: 90%;
    border-radius: 10px;
    box-shadow: 0 10px 30px rgba(0,0,0,0.5);
}

.modal-info {
    color: white;
    text-align: center;
    margin-top: 15px;
    font-size: 18px;
}

.close {
    position: absolute;
    top: 20px;
    right: 30px;
    color: white;
    font-size: 40px;
    font-weight: bold;
    cursor: pointer;
}

@media print {
    body {
        background: white;
        padding: 0;
    }
    .container {
        box-shadow: none;
        border-radius: 0;
    }
    .actions {
        display: none;
    }
    .modal {
        display: none !important;
    }
}

/* Фильтры */
.filters {
    display: flex;
    gap: 15px;
    margin-bottom: 20px;
    flex-wrap: wrap;
}

.filter-group {
    display: flex;
    align-items: center;
    gap: 8px;
}

.filter-select {
    padding: 8px 12px;
    border: 1px solid #ddd;
    border-radius: 6px;
    background: white;
    min-width: 150px;
}

.filter-btn {
    padding: 8px 16px;
    background: #4a6fa5;
    color: white;
    border: none;
    border-radius: 6px;
    cursor: pointer;
}

.filter-btn:hover {
    background: #3a5a8a;
}

/* Скроллбар для таблицы */
.table-wrapper::-webkit-scrollbar {
    width: 10px;
    height: 10px;
}

.table-wrapper::-webkit-scrollbar-track {
    background: #f1f1f1;
    border-radius: 5px;
}

.table-wrapper::-webkit-scrollbar-thumb {
    background: #888;
    border-radius: 5px;
}

.table-wrapper::-webkit-scrollbar-thumb:hover {
    background: #555;
}

/* Адаптивность */
@media (max-width: 1200px) {
    .container {
        max-width: 100%;
    }
    
    .stats-grid {
        grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
    }
}

@media (max-width: 768px) {
    .header h1 {
        font-size: 1.8rem;
    }
    
    .summary-cards {
        grid-template-columns: repeat(2, 1fr);
    }
    
    .filters {
        flex-direction: column;
        align-items: stretch;
    }
    
    .filter-group {
        justify-content: space-between;
    }
    
    .actions {
        flex-direction: column;
        align-items: stretch;
    }
    
    .btn {
        width: 100%;
        justify-content: center;
    }
}

@media (max-width: 480px) {
    .summary-cards {
        grid-template-columns: 1fr;
    }
    
    .card {
        flex-direction: column;
        text-align: center;
    }
}'''
    
    def _get_javascript_code(self) -> str:
        """Возвращает JavaScript код"""
        return '''
        // Функция для сохранения как PDF
        function saveAsPDF() {
            const element = document.querySelector('.container');
            
            const opt = {
                margin: [10, 10, 10, 10],
                filename: 'face_recognition_report.pdf',
                image: { type: 'jpeg', quality: 0.98 },
                html2canvas: { 
                    scale: 2,
                    useCORS: true,
                    logging: true,
                    windowWidth: document.documentElement.offsetWidth,
                    windowHeight: document.documentElement.offsetHeight
                },
                jsPDF: { 
                    unit: 'mm', 
                    format: 'a4', 
                    orientation: 'landscape' 
                }
            };
            
            if (typeof html2pdf !== 'undefined') {
                html2pdf().set(opt).from(element).save();
                showNotification('PDF создается...', 'info');
            } else {
                const script = document.createElement('script');
                script.src = 'https://cdnjs.cloudflare.com/ajax/libs/html2pdf.js/0.10.1/html2pdf.bundle.min.js';
                script.onload = function() {
                    html2pdf().set(opt).from(element).save();
                    showNotification('PDF создается...', 'info');
                };
                document.head.appendChild(script);
                
                showNotification('Загружаем библиотеку для создания PDF...', 'warning');
            }
        }
        
        // Функция для экспорта в Excel
        function exportToExcel() {
            const table = document.getElementById('dataTable');
            const rows = table.querySelectorAll('tr');
            let csv = [];
            
            rows.forEach(row => {
                const cells = row.querySelectorAll('td, th');
                const rowData = [];
                
                cells.forEach(cell => {
                    // Пропускаем колонку с фото
                    if (!cell.classList.contains('image-cell') && !cell.classList.contains('no-image')) {
                        // Очищаем текст от HTML тегов
                        let cellText = cell.innerText || cell.textContent || '';
                        // Экранируем кавычки
                        cellText = cellText.replace(/"/g, '""');
                        // Добавляем кавычки если есть запятые или кавычки
                        if (cellText.includes(',') || cellText.includes('"')) {
                            cellText = '"' + cellText + '"';
                        }
                        rowData.push(cellText);
                    }
                });
                
                if (rowData.length > 0) {
                    csv.push(rowData.join(','));
                }
            });
            
            const csvContent = 'data:text/csv;charset=utf-8,' + csv.join('\\n');
            const encodedUri = encodeURI(csvContent);
            const link = document.createElement('a');
            link.setAttribute('href', encodedUri);
            link.setAttribute('download', 'face_recognition_data.csv');
            document.body.appendChild(link);
            link.click();
            document.body.removeChild(link);
            
            showNotification('Excel файл загружается...', 'success');
        }
        
        // Функции для фильтрации таблицы
        function filterTable() {
            const companyFilter = document.getElementById('companyFilter').value;
            const eventFilter = document.getElementById('eventFilter').value;
            const listFilter = document.getElementById('listFilter').value;
            
            const rows = document.querySelectorAll('#dataTable tbody tr');
            let visibleCount = 0;
            
            rows.forEach(row => {
                const companyCell = row.querySelector('td:nth-child(9)'); // 9-я колонка - компания
                const eventTypeCell = row.querySelector('td:nth-child(10)'); // 10-я колонка - тип
                const listStatusCell = row.querySelector('td:nth-child(11)'); // 11-я колонка - список
                
                if (!companyCell || !eventTypeCell || !listStatusCell) {
                    row.style.display = 'none';
                    return;
                }
                
                const company = companyCell.textContent.trim();
                const eventTypeText = eventTypeCell.textContent.trim();
                const listStatusText = listStatusCell.textContent.trim();
                
                const eventType = eventTypeText.includes('Распознавание') ? '1' : '2';
                const listStatus = listStatusText.includes('В списке') ? '1' : '0';
                
                let show = true;
                
                if (companyFilter && company !== companyFilter) {
                    show = false;
                }
                
                if (eventFilter && eventType !== eventFilter) {
                    show = false;
                }
                
                if (listFilter && listStatus !== listFilter) {
                    show = false;
                }
                
                if (show) {
                    row.style.display = '';
                    visibleCount++;
                } else {
                    row.style.display = 'none';
                }
            });
            
            // Показываем количество отфильтрованных записей
            showNotification(`Отфильтровано: ${visibleCount} записей`, 'info');
        }
        
        function resetFilters() {
            document.getElementById('companyFilter').value = '';
            document.getElementById('eventFilter').value = '';
            document.getElementById('listFilter').value = '';
            filterTable();
            showNotification('Фильтры сброшены', 'info');
        }
        
        // Функции для предпросмотра фото
        function showImagePreview(imageSrc, userName, width, height) {
            const modal = document.getElementById('imageModal');
            const modalImg = document.getElementById('modalImage');
            const modalInfo = document.getElementById('modalInfo');
            
            modal.style.display = 'flex';
            modalImg.src = imageSrc;
            modalInfo.innerHTML = `
                <div style="margin-bottom: 10px;">
                    <strong>Пользователь:</strong> ${userName}
                </div>
                <div>
                    <strong>Размер:</strong> ${width}×${height} пикселей
                </div>
            `;
        }
        
        function closeModal() {
            document.getElementById('imageModal').style.display = 'none';
        }
        
        // Закрытие модального окна при клике на крестик
        document.querySelector('.close').onclick = closeModal;
        
        // Закрытие модального окна при клике вне изображения
        window.onclick = function(event) {
            const modal = document.getElementById('imageModal');
            if (event.target === modal) {
                closeModal();
            }
        }
        
        // Сортировка таблицы
        document.addEventListener('DOMContentLoaded', function() {
            const headers = document.querySelectorAll('th');
            headers.forEach((header, index) => {
                if (index !== 11) { // Не сортируем колонку с фото
                    header.style.cursor = 'pointer';
                    header.addEventListener('click', () => sortTable(index));
                }
            });
        });
        
        let sortDirection = true;
        
        function sortTable(column) {
            const table = document.querySelector('table');
            const tbody = table.querySelector('tbody');
            const rows = Array.from(tbody.querySelectorAll('tr'));
            
            rows.sort((a, b) => {
                const aCell = a.cells[column];
                const bCell = b.cells[column];
                
                if (!aCell || !bCell) return 0;
                
                const aText = aCell.textContent.trim();
                const bText = bCell.textContent.trim();
                
                const aNum = parseFloat(aText.replace('%', '').replace(',', '.'));
                const bNum = parseFloat(bText.replace('%', '').replace(',', '.'));
                
                if (!isNaN(aNum) && !isNaN(bNum)) {
                    return sortDirection ? aNum - bNum : bNum - aNum;
                }
                
                return sortDirection 
                    ? aText.localeCompare(bText, 'ru', { numeric: true })
                    : bText.localeCompare(aText, 'ru', { numeric: true });
            });
            
            // Обновляем номера строк
            rows.forEach((row, index) => {
                const numCell = row.cells[0];
                if (numCell) {
                    numCell.textContent = index + 1;
                }
                tbody.appendChild(row);
            });
            
            sortDirection = !sortDirection;
            showNotification(`Таблица отсортирована по колонке ${column + 1}`, 'info');
        }
        
        // Создание графиков
        document.addEventListener('DOMContentLoaded', function() {
            createCharts();
        });
        
        function createCharts() {
            // График распределения по компаниям
            const companyCtx = document.getElementById('companyChart');
            if (companyCtx && companyLabels && companyData) {
                try {
                    new Chart(companyCtx, {
                        type: 'bar',
                        data: {
                            labels: companyLabels,
                            datasets: [{
                                label: 'Количество записей',
                                data: companyData,
                                backgroundColor: 'rgba(54, 162, 235, 0.5)',
                                borderColor: 'rgba(54, 162, 235, 1)',
                                borderWidth: 1
                            }]
                        },
                        options: {
                            responsive: true,
                            plugins: {
                                title: {
                                    display: true,
                                    text: 'Распределение по компаниям (топ 15)'
                                },
                                legend: {
                                    display: false
                                }
                            },
                            scales: {
                                y: {
                                    beginAtZero: true,
                                    ticks: {
                                        stepSize: 1
                                    }
                                }
                            }
                        }
                    });
                } catch (e) {
                    console.error('Error creating company chart:', e);
                }
            }
            
            // График распределения по полу
            const genderCtx = document.getElementById('genderChart');
            if (genderCtx && genderLabels && genderData) {
                try {
                    new Chart(genderCtx, {
                        type: 'pie',
                        data: {
                            labels: genderLabels,
                            datasets: [{
                                data: genderData,
                                backgroundColor: [
                                    'rgba(255, 99, 132, 0.5)',
                                    'rgba(54, 162, 235, 0.5)',
                                    'rgba(255, 206, 86, 0.5)'
                                ],
                                borderColor: [
                                    'rgba(255, 99, 132, 1)',
                                    'rgba(54, 162, 235, 1)',
                                    'rgba(255, 206, 86, 1)'
                                ],
                                borderWidth: 1
                            }]
                        },
                        options: {
                            responsive: true,
                            plugins: {
                                title: {
                                    display: true,
                                    text: 'Распределение по полу'
                                }
                            }
                        }
                    });
                } catch (e) {
                    console.error('Error creating gender chart:', e);
                }
            }
        }
        
        // Функция для показа уведомлений
        function showNotification(message, type = 'info') {
            // Удаляем предыдущие уведомления
            const oldNotification = document.querySelector('.notification');
            if (oldNotification) {
                oldNotification.remove();
            }
            
            const notification = document.createElement('div');
            notification.className = `notification notification-${type}`;
            notification.innerHTML = `
                <span>${message}</span>
                <button onclick="this.parentElement.remove()">&times;</button>
            `;
            
            // Стили для уведомления
            notification.style.cssText = `
                position: fixed;
                top: 20px;
                right: 20px;
                padding: 15px 20px;
                background: ${type === 'success' ? '#4CAF50' : type === 'warning' ? '#FF9800' : '#2196F3'};
                color: white;
                border-radius: 8px;
                box-shadow: 0 4px 12px rgba(0,0,0,0.2);
                display: flex;
                align-items: center;
                justify-content: space-between;
                gap: 15px;
                z-index: 10000;
                animation: slideIn 0.3s ease;
                max-width: 400px;
            `;
            
            document.body.appendChild(notification);
            
            // Автоматическое удаление через 3 секунды
            setTimeout(() => {
                if (notification.parentElement) {
                    notification.remove();
                }
            }, 3000);
        }
        
        // Анимация для уведомления
        const style = document.createElement('style');
        style.textContent = `
            @keyframes slideIn {
                from { transform: translateX(100%); opacity: 0; }
                to { transform: translateX(0); opacity: 1; }
            }
        `;
        document.head.appendChild(style);
        '''
    
    def _generate_stats_html(self, stats: Dict, metrics) -> str:
        """Генерация HTML статистики"""
        # Статистика по компаниям (топ 5)
        top_companies = sorted(stats['by_company'].items(), key=lambda x: x[1], reverse=True)[:5]
        companies_html = ""
        for company, count in top_companies:
            percentage = (count / stats['total_records']) * 100 if stats['total_records'] > 0 else 0
            companies_html += f'''
            <div class="stat-item">
                <span class="stat-label">{html.escape(str(company))}</span>
                <span class="stat-value">{count} ({percentage:.1f}%)</span>
            </div>
            '''
        
        # Статистика по полу
        gender_html = ""
        for gender, count in stats['by_gender'].items():
            percentage = (count / stats['total_records']) * 100 if stats['total_records'] > 0 else 0
            gender_html += f'''
            <div class="stat-item">
                <span class="stat-label">{html.escape(str(gender))}</span>
                <span class="stat-value">{count} ({percentage:.1f}%)</span>
            </div>
            '''
        
        # Статистика по возрастным группам
        age_html = ""
        for group, count in stats['by_age_group'].items():
            percentage = (count / stats['total_records']) * 100 if stats['total_records'] > 0 else 0
            age_html += f'''
            <div class="stat-item">
                <span class="stat-label">{html.escape(str(group))}</span>
                <span class="stat-value">{count} ({percentage:.1f}%)</span>
            </div>
            '''
        
        # Статистика по оценкам
        score_html = ""
        for score_range, count in stats['score_distribution'].items():
            percentage = (count / stats['total_records']) * 100 if stats['total_records'] > 0 else 0
            score_html += f'''
            <div class="stat-item">
                <span class="stat-label">{html.escape(str(score_range))}</span>
                <span class="stat-value">{count} ({percentage:.1f}%)</span>
            </div>
            '''
        
        return f'''
        <div class="stats-grid">
            <div class="stat-box">
                <h3><i class="fas fa-building"></i> По компаниям (топ 5)</h3>
                {companies_html}
                <div class="chart-container">
                    <canvas id="companyChart"></canvas>
                </div>
            </div>
            
            <div class="stat-box">
                <h3><i class="fas fa-venus-mars"></i> По полу</h3>
                {gender_html}
                <div class="chart-container">
                    <canvas id="genderChart"></canvas>
                </div>
            </div>
            
            <div class="stat-box">
                <h3><i class="fas fa-users"></i> По возрасту</h3>
                {age_html}
            </div>
            
            <div class="stat-box">
                <h3><i class="fas fa-chart-pie"></i> Общая статистика</h3>
                <div class="stat-item">
                    <span class="stat-label">Всего записей</span>
                    <span class="stat-value">{stats['total_records']:,}</span>
                </div>
                <div class="stat-item">
                    <span class="stat-label">С фото</span>
                    <span class="stat-value">{stats['with_images']:,}</span>
                </div>
                <div class="stat-item">
                    <span class="stat-label">Без фото</span>
                    <span class="stat-value">{stats['without_images']:,}</span>
                </div>
                <div class="stat-item">
                    <span class="stat-label">Распознавания</span>
                    <span class="stat-value">{stats['by_event_type'].get('1', 0):,}</span>
                </div>
                <div class="stat-item">
                    <span class="stat-label">События</span>
                    <span class="stat-value">{stats['by_event_type'].get('2', 0):,}</span>
                </div>
            </div>
            
            <div class="stat-box">
                <h3><i class="fas fa-percentage"></i> По совпадению</h3>
                {score_html}
            </div>
            
            <div class="stat-box">
                <h3><i class="fas fa-trophy"></i> Топ пользователей</h3>
                {self._generate_top_list_html(stats['top_users'])}
            </div>
            
            <div class="stat-box">
                <h3><i class="fas fa-microchip"></i> Топ устройств</h3>
                {self._generate_top_list_html(stats['top_devices'], is_device=True)}
            </div>
        </div>
        '''
    
    def _generate_top_list_html(self, items: Dict, is_device: bool = False) -> str:
        """Генерация HTML для топ-списка"""
        result = ""
        for i, (item, count) in enumerate(list(items.items())[:10], 1):
            item_str = str(item)
            display_name = html.escape(item_str[:30]) + ('...' if len(item_str) > 30 else '')
            icon = "📱" if is_device else "👤"
            result += f'''
            <div class="stat-item">
                <span class="stat-label">
                    {i}. {icon} {display_name}
                </span>
                <span class="stat-value">{count}</span>
            </div>
            '''
        return result
    
    def _generate_filter_options(self, field: str, records: List) -> str:
        """Генерация опций для фильтров"""
        values = set()
        for record in records:
            if field == 'company':
                value = str(record.company_id)
                if value and value != 'Н/Д':
                    values.add(value)
        
        options = ""
        for value in sorted(values)[:50]:  # Ограничиваем 50 опциями
            options += f'<option value="{html.escape(value)}">{html.escape(value)}</option>\n'
        
        if len(values) > 50:
            options += f'<option value="">... и еще {len(values) - 50} компаний</option>\n'
        
        return options
    
    def generate_pdf_report(self, records: List, metrics) -> Optional[str]:
        """Генерация PDF отчета"""
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            
            pdf_path = os.path.join(self.reports_dir, Config.PDF_REPORT)
            
            # Создаем документ
            doc = SimpleDocTemplate(
                pdf_path,
                pagesize=landscape(A4),
                rightMargin=1.5*cm,
                leftMargin=1.5*cm,
                topMargin=2*cm,
                bottomMargin=2*cm
            )
            
            elements = []
            styles = getSampleStyleSheet()
            
            # Создаем свой стиль для заголовков
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                alignment=1,  # Выравнивание по центру
                spaceAfter=20,
                textColor=colors.HexColor('#2c3e50')
            )
            
            # Заголовок
            title = Paragraph("Отчет по распознаванию лиц", title_style)
            elements.append(title)
            elements.append(Spacer(1, 12))
            
            # Информация
            info_style = ParagraphStyle(
                'InfoStyle',
                parent=styles['Normal'],
                fontSize=10,
                alignment=0,
                spaceAfter=10
            )
            
            current_time = datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S")
            info_text = f"""
            <b>Дата создания:</b> {current_time}<br/>
            <b>Всего записей:</b> {metrics.total_records:,}<br/>
            <b>Успешных фото:</b> {metrics.valid_images:,}<br/>
            <b>Ошибок загрузки:</b> {metrics.failed_images:,}<br/>
            <b>Время обработки:</b> {metrics.elapsed_time:.1f} сек.<br/>
            <b>Уникальных пользователей:</b> {len(metrics.unique_users):,}
            """
            info = Paragraph(info_text, info_style)
            elements.append(info)
            elements.append(Spacer(1, 20))
            
            # Создаем таблицу (ограничиваем количество записей для PDF)
            max_records_for_pdf = min(len(records), 200)
            data = []
            
            # Заголовки таблицы
            headers = [
                "№", "Время", "Устройство", "Пользователь", "Пол", 
                "Возраст", "Совпадение", "Компания", "Тип"
            ]
            data.append(headers)
            
            # Данные
            for i, record in enumerate(records[:max_records_for_pdf]):
                row = [
                    str(i + 1),
                    record.timestamp[:19] if len(record.timestamp) > 19 else record.timestamp,
                    str(record.device_id)[:15] + ("..." if len(str(record.device_id)) > 15 else ""),
                    str(record.user_name)[:20] + ("..." if len(str(record.user_name)) > 20 else ""),
                    record.gender[:10],
                    record.age,
                    record.score,
                    str(record.company_id)[:15] + ("..." if len(str(record.company_id)) > 15 else ""),
                    "Распозн." if record.event_type == "1" else "Событие"
                ]
                data.append(row)
            
            # Создаем таблицу
            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            elements.append(table)
            elements.append(Spacer(1, 20))
            
            # Примечание
            note_style = ParagraphStyle(
                'NoteStyle',
                parent=styles['Italic'],
                fontSize=8,
                alignment=0,
                textColor=colors.grey
            )
            
            note = Paragraph(
                f"<i>Отчет содержит {len(records):,} записей. В PDF показаны первые {max_records_for_pdf} записей. "
                f"Для полных данных откройте HTML отчет.</i>",
                note_style
            )
            elements.append(note)
            
            # Строим PDF
            doc.build(elements)
            
            print(f"✅ PDF отчет создан: {pdf_path}")
            return pdf_path
            
        except ImportError as e:
            print(f"⚠️  ReportLab не установлен. PDF отчет не будет создан.")
            print(f"   Ошибка: {e}")
            return None
        except Exception as e:
            print(f"❌ Ошибка создания PDF: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def generate_excel_report(self, records: List) -> Optional[str]:
        """Генерация Excel отчета"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Распознавание лиц"
            
            # Заголовки
            headers = [
                "№", "Время", "Устройство", "Пользователь", "Пол", 
                "Возраст", "Совпадение %", "ID Лица", "ID Компании",
                "Тип события", "Статус списка", "IP Адрес",
                "URL Изображения", "Хэш изображения", "Файл фото"
            ]
            
            # Стили для заголовков
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(
                start_color="2C3E50",
                end_color="2C3E50",
                fill_type="solid"
            )
            alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Записываем заголовки
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
                cell.border = thin_border
            
            # Данные
            print(f"📊 Запись {len(records)} записей в Excel...")
            for row_idx, record in enumerate(records, 2):
                data = record.to_dict()
                row_data = [
                    row_idx - 1,  # №
                    data['timestamp'],
                    str(data['device_id']),
                    str(data['user_name']),
                    data['gender'],
                    data['age'],
                    data['score'],
                    str(data['face_id']),
                    str(data['company_id']),
                    "Распознавание" if data['event_type'] == "1" else "Событие",
                    "В списке" if data['user_list'] == "1" else "Не в списке",
                    str(data['ip_address']),
                    data['image_url'][:200],  # Ограничиваем длину URL
                    data['image_hash'],
                    os.path.basename(data['image_path']) if data['image_path'] else ""
                ]
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
            
            # Автоширина столбцов
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        cell_value = str(cell.value) if cell.value is not None else ""
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Замораживаем заголовки
            ws.freeze_panes = 'A2'
            
            # Добавляем фильтр
            ws.auto_filter.ref = ws.dimensions
            
            # Сохранение
            excel_path = os.path.join(self.reports_dir, Config.EXCEL_REPORT)
            wb.save(excel_path)
            
            print(f"✅ Excel отчет создан: {excel_path}")
            return excel_path
            
        except ImportError as e:
            print(f"⚠️  OpenPyXL не установлен. Excel отчет не будет создан.")
            print(f"   Ошибка: {e}")
            return None
        except Exception as e:
            print(f"❌ Ошибка создания Excel отчета: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def generate_summary_report(self, metrics, records: List) -> str:
        """Генерация JSON отчета с метаданными"""
        print("📊 Создание JSON отчета...")
        
        # Используем отложенный импорт
        from core.statistics import StatisticsAnalyzer
        
        stats = StatisticsAnalyzer.analyze(records)
        
        # Преобразуем статистику для JSON
        by_company = {str(k): v for k, v in stats['by_company'].items()}
        by_gender = {str(k): v for k, v in stats['by_gender'].items()}
        by_age_group = {str(k): v for k, v in stats['by_age_group'].items()}
        by_event_type = {str(k): v for k, v in stats['by_event_type'].items()}
        top_users = {str(k): v for k, v in stats['top_users'].items()}
        top_devices = {str(k): v for k, v in stats['top_devices'].items()}
        
        summary = {
            "metadata": {
                "generated_at": datetime.datetime.now().isoformat(),
                "version": f"FaceRecognitionAnalytics v{Config.VERSION}",
                "processing_time_seconds": metrics.elapsed_time,
                "total_records": metrics.total_records,
                "valid_images": metrics.valid_images,
                "failed_images": metrics.failed_images,
                "success_rate_percent": metrics.success_rate,
                "cached_images": metrics.cached_images,
                "json_errors": metrics.json_errors,
                "network_errors": metrics.network_errors,
                "timeout_errors": metrics.timeout_errors,
                "duplicate_records": metrics.duplicate_records,
                "unique_users": len(metrics.unique_users),
                "unique_devices": len(metrics.unique_devices),
                "unique_companies": len(metrics.unique_companies),
                "unique_ips": len(metrics.unique_ips),
                "records_per_second": metrics.total_records / metrics.elapsed_time if metrics.elapsed_time > 0 else 0,
            },
            "statistics": {
                "by_company": by_company,
                "by_gender": by_gender,
                "by_age_group": by_age_group,
                "by_event_type": by_event_type,
                "score_distribution": stats['score_distribution'],
                "with_images": stats['with_images'],
                "without_images": stats['without_images'],
                "top_users": top_users,
                "top_devices": top_devices
            },
            "files": {
                "html_report": Config.HTML_REPORT,
                "pdf_report": Config.PDF_REPORT,
                "excel_report": Config.EXCEL_REPORT,
                "images_count": len([f for f in os.listdir(self.images_dir) if f.endswith('.jpg')])
            }
        }
        
        summary_path = os.path.join(self.reports_dir, Config.SUMMARY_REPORT)
        with open(summary_path, 'w', encoding='utf-8') as f:
            json.dump(summary, f, indent=2, ensure_ascii=False)
        
        print(f"✅ JSON отчет создан: {summary_path}")
        return summary_path
    
    def _create_fallback_html_report(self, records: List, metrics) -> str:
        """Создание простого HTML отчета в случае ошибки"""
        try:
            html_content = f'''<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <title>Отчет по распознаванию лиц (упрощенный)</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; }}
        table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background-color: #f2f2f2; }}
        tr:nth-child(even) {{ background-color: #f9f9f9; }}
    </style>
</head>
<body>
    <h1>Отчет по распознаванию лиц</h1>
    <p><strong>Дата создания:</strong> {datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S")}</p>
    <p><strong>Всего записей:</strong> {metrics.total_records:,}</p>
    <p><strong>Успешных фото:</strong> {metrics.valid_images:,}</p>
    <p><strong>Ошибок загрузки:</strong> {metrics.failed_images:,}</p>
    
    <h2>Данные ({len(records)} записей)</h2>
    <table>
        <thead>
            <tr>
                <th>№</th>
                <th>Время</th>
                <th>Пользователь</th>
                <th>Пол</th>
                <th>Возраст</th>
                <th>Совпадение</th>
                <th>Компания</th>
                <th>Тип</th>
            </tr>
        </thead>
        <tbody>
'''
            
            for i, record in enumerate(records[:1000]):  # Ограничиваем 1000 записей
                html_content += f'''
            <tr>
                <td>{i + 1}</td>
                <td>{html.escape(record.timestamp)}</td>
                <td>{html.escape(str(record.user_name))}</td>
                <td>{html.escape(record.gender)}</td>
                <td>{html.escape(record.age)}</td>
                <td>{html.escape(record.score)}</td>
                <td>{html.escape(str(record.company_id))}</td>
                <td>{'Распознавание' if record.event_type == '1' else 'Событие'}</td>
            </tr>
'''
            
            html_content += f'''
        </tbody>
    </table>
    
    <p><em>Отчет создан в упрощенном формате из-за ошибки генерации основного отчета.</em></p>
    <p><em>Показаны первые {min(len(records), 1000)} записей из {len(records)}.</em></p>
</body>
</html>'''
            
            report_path = os.path.join(self.reports_dir, "fallback_report.html")
            with open(report_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            print(f"⚠️  Создан упрощенный HTML отчет: {report_path}")
            return report_path
            
        except Exception as e:
            print(f"❌ Не удалось создать даже упрощенный отчет: {e}")
            # Создаем минимальный файл
            report_path = os.path.join(self.reports_dir, "error_report.txt")
            with open(report_path, 'w', encoding='utf-8') as f:
                f.write(f"Ошибка создания отчета: {e}\n")
                f.write(f"Всего записей: {metrics.total_records}\n")
                f.write(f"Дата: {datetime.datetime.now()}\n")
            return report_path